from __future__ import annotations

import argparse
import warnings
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, func, select

from app.database import SessionLocal
from app.models import ServiceCatalog


DEFAULT_MINIMUM_DEPARTMENT_TICKET_COUNT = 35

INVALID_VALUES = {
    "",
    "string",
    "null",
    "none",
    "nan",
    "not assigned",
}

REQUIRED_HEADERS = {
    "department",
    "category",
    "subcategory",
}


warnings.filterwarnings(
    "ignore",
    message=(
        "Workbook contains no default style.*"
    ),
    category=UserWarning,
)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Excel dosyasındaki departman, kategori "
            "ve alt kategori ilişkilerini "
            "service_catalog tablosuna aktarır."
        )
    )

    parser.add_argument(
        "excel_path",
        type=Path,
        help="Service Desk Excel dosyasının yolu.",
    )

    parser.add_argument(
        "--minimum-count",
        type=int,
        default=(
            DEFAULT_MINIMUM_DEPARTMENT_TICKET_COUNT
        ),
        help=(
            "Bir departmanın kataloğa alınması için "
            "gereken minimum ticket sayısı. "
            "Varsayılan: 35"
        ),
    )

    return parser.parse_args()


def clean_value(
    value,
) -> str | None:
    if value is None:
        return None

    cleaned_value = " ".join(
        str(value).strip().split()
    )

    if (
        cleaned_value.casefold()
        in INVALID_VALUES
    ):
        return None

    return cleaned_value


def find_catalog_worksheet(
    workbook,
):
    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=20,
                values_only=True,
            ),
            start=1,
        ):
            normalized_headers = {
                str(value).strip().casefold(): index
                for index, value in enumerate(row)
                if value is not None
            }

            if REQUIRED_HEADERS.issubset(
                normalized_headers
            ):
                return (
                    worksheet,
                    row_number,
                    normalized_headers,
                )

    raise RuntimeError(
        "Department, Category ve Subcategory "
        "sütunlarını içeren çalışma sayfası "
        "bulunamadı."
    )


def read_catalog_relations(
    excel_path: Path,
    minimum_count: int,
):
    workbook = load_workbook(
        excel_path,
        read_only=True,
        data_only=True,
    )

    try:
        (
            worksheet,
            header_row_number,
            header_indexes,
        ) = find_catalog_worksheet(
            workbook
        )

        department_index = (
            header_indexes["department"]
        )

        category_index = (
            header_indexes["category"]
        )

        subcategory_index = (
            header_indexes["subcategory"]
        )

        department_counts = Counter()
        valid_rows = []

        for row in worksheet.iter_rows(
            min_row=header_row_number + 1,
            values_only=True,
        ):
            department = clean_value(
                row[department_index]
            )

            category = clean_value(
                row[category_index]
            )

            subcategory = clean_value(
                row[subcategory_index]
            )

            if department:
                department_counts[
                    department
                ] += 1

            if (
                department
                and category
                and subcategory
            ):
                valid_rows.append(
                    (
                        department,
                        category,
                        subcategory,
                    )
                )

        selected_departments = {
            department
            for department, ticket_count
            in department_counts.items()
            if ticket_count >= minimum_count
        }

        catalog_relations = {}

        for (
            department,
            category,
            subcategory,
        ) in valid_rows:
            if (
                department
                not in selected_departments
            ):
                continue

            normalized_key = (
                department.casefold(),
                category.casefold(),
                subcategory.casefold(),
            )

            catalog_relations.setdefault(
                normalized_key,
                (
                    department,
                    category,
                    subcategory,
                ),
            )

        sorted_relations = sorted(
            catalog_relations.values(),
            key=lambda item: (
                item[0].casefold(),
                item[1].casefold(),
                item[2].casefold(),
            ),
        )

        if not selected_departments:
            raise RuntimeError(
                "Belirtilen minimum ticket "
                "sayısına uygun departman bulunamadı."
            )

        if not sorted_relations:
            raise RuntimeError(
                "Aktarılabilecek geçerli katalog "
                "bağlantısı bulunamadı."
            )

        return {
            "worksheet_name": worksheet.title,
            "department_counts": (
                department_counts
            ),
            "selected_departments": (
                selected_departments
            ),
            "relations": sorted_relations,
        }
    finally:
        workbook.close()


def replace_service_catalog(
    relations: list[
        tuple[str, str, str]
    ],
) -> dict[str, int]:
    with SessionLocal() as database:
        try:
            database.execute(
                delete(ServiceCatalog)
            )

            database.add_all(
                [
                    ServiceCatalog(
                        department=department,
                        category=category,
                        subcategory=subcategory,
                    )
                    for (
                        department,
                        category,
                        subcategory,
                    ) in relations
                ]
            )

            database.commit()
        except Exception:
            database.rollback()
            raise

        total_rows = int(
            database.scalar(
                select(
                    func.count(
                        ServiceCatalog.catalog_id
                    )
                )
            )
            or 0
        )

        department_count = int(
            database.scalar(
                select(
                    func.count(
                        func.distinct(
                            ServiceCatalog.department
                        )
                    )
                )
            )
            or 0
        )

        category_count = int(
            database.scalar(
                select(
                    func.count(
                        func.distinct(
                            ServiceCatalog.category
                        )
                    )
                )
            )
            or 0
        )

        subcategory_count = int(
            database.scalar(
                select(
                    func.count(
                        func.distinct(
                            ServiceCatalog.subcategory
                        )
                    )
                )
            )
            or 0
        )

    return {
        "total_rows": total_rows,
        "department_count": department_count,
        "category_count": category_count,
        "subcategory_count": subcategory_count,
    }


def main() -> None:
    arguments = parse_arguments()

    excel_path = (
        arguments.excel_path
        .expanduser()
        .resolve()
    )

    minimum_count = (
        arguments.minimum_count
    )

    if minimum_count < 1:
        raise ValueError(
            "Minimum ticket sayısı en az 1 "
            "olmalıdır."
        )

    if not excel_path.exists():
        raise FileNotFoundError(
            f"Excel dosyası bulunamadı: "
            f"{excel_path}"
        )

    if not excel_path.is_file():
        raise ValueError(
            f"Belirtilen yol bir dosya değil: "
            f"{excel_path}"
        )

    catalog_data = read_catalog_relations(
        excel_path=excel_path,
        minimum_count=minimum_count,
    )

    database_counts = (
        replace_service_catalog(
            relations=catalog_data[
                "relations"
            ],
        )
    )

    selected_departments = sorted(
        catalog_data[
            "selected_departments"
        ],
        key=lambda department: (
            -catalog_data[
                "department_counts"
            ][department],
            department.casefold(),
        ),
    )

    print()
    print(
        "Excel:",
        excel_path,
    )

    print(
        "Sayfa:",
        catalog_data["worksheet_name"],
    )

    print(
        "Minimum departman ticket sayısı:",
        minimum_count,
    )

    print(
        "Seçilen departman sayısı:",
        database_counts[
            "department_count"
        ],
    )

    print(
        "Eklenen katalog bağlantısı:",
        database_counts[
            "total_rows"
        ],
    )

    print(
        "Kategori sayısı:",
        database_counts[
            "category_count"
        ],
    )

    print(
        "Alt kategori sayısı:",
        database_counts[
            "subcategory_count"
        ],
    )

    print()
    print("SEÇİLEN DEPARTMANLAR:")

    for department in selected_departments:
        print(
            f"- {department}: "
            f"{catalog_data['department_counts'][department]}"
        )


if __name__ == "__main__":
    main()